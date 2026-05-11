"""Экспорт payroll-отчёта в XLSX / PDF + универсальный CSV/PDF снапшот статистики."""

from __future__ import annotations

import csv
from io import BytesIO, StringIO
from typing import Any

from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Font


def payroll_to_xlsx(rows: list[dict[str, Any]], *, period_label: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Payroll"
    ws.append(["Payroll report", period_label])
    ws.append([])
    header = [
        "Master",
        "Revenue",
        "Completed",
        "Payroll %",
        "Payroll amount",
    ]
    ws.append(header)
    for c in ws[3]:
        c.font = Font(bold=True)
    total_rev = 0.0
    total_pay = 0.0
    for r in rows:
        rev = float(r.get("revenue", "0") or 0)
        pay = float(r.get("payroll_amount", "0") or 0)
        total_rev += rev
        total_pay += pay
        ws.append(
            [
                r.get("display_name", ""),
                r.get("revenue", ""),
                r.get("completed_bookings", ""),
                r.get("payroll_percent", ""),
                r.get("payroll_amount", ""),
            ]
        )
    ws.append([])
    ws.append(["Total", "", "", "", f"{total_pay:.2f}"])
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def payroll_to_pdf(rows: list[dict[str, Any]], *, period_label: str, title: str = "Payroll") -> bytes:
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 10, title, ln=True)
    pdf.set_font("Helvetica", size=10)
    pdf.cell(0, 8, period_label, ln=True)
    pdf.ln(4)
    pdf.set_font("Helvetica", "B", 9)
    col_w = 38
    for h in ["Master", "Revenue", "Done", "%", "Payroll"]:
        pdf.cell(col_w, 7, h[:18], border=1)
    pdf.ln()
    pdf.set_font("Helvetica", size=9)
    total_pay = 0.0
    for r in rows:
        pay = float(r.get("payroll_amount", "0") or 0)
        total_pay += pay
        pdf.cell(col_w, 7, str(r.get("display_name", ""))[:22], border=1)
        pdf.cell(col_w, 7, str(r.get("revenue", ""))[:12], border=1)
        pdf.cell(col_w, 7, str(r.get("completed_bookings", ""))[:8], border=1)
        pdf.cell(col_w, 7, str(r.get("payroll_percent", ""))[:8], border=1)
        pdf.cell(col_w, 7, str(r.get("payroll_amount", ""))[:12], border=1)
        pdf.ln()
    pdf.ln(2)
    pdf.set_font("Helvetica", "B", 10)
    pdf.cell(0, 8, f"Total payroll: {total_pay:.2f}", ln=True)
    out = pdf.output(dest="S")
    if isinstance(out, str):
        return out.encode("latin-1")
    return bytes(out)


def stats_snapshot_to_csv(
    *,
    period_label: str,
    kpi: dict[str, Any],
    revenue_trend: list[dict[str, Any]],
    masters: list[dict[str, Any]],
    services_top: list[dict[str, Any]],
) -> bytes:
    """CSV-снапшот статистики: KPI + тренд + мастера + услуги."""
    buf = StringIO()
    w = csv.writer(buf)
    w.writerow(["Statistics report", period_label])
    w.writerow([])
    w.writerow(["KPI"])
    for k, v in kpi.items():
        w.writerow([k, v])
    w.writerow([])
    w.writerow(["Revenue trend"])
    w.writerow(["date", "revenue", "bookings_count"])
    for r in revenue_trend:
        w.writerow([r.get("date", ""), r.get("revenue", ""), r.get("bookings_count", "")])
    w.writerow([])
    w.writerow(["Masters"])
    w.writerow(["display_name", "revenue", "completed_bookings", "avg_check", "utilization_pct"])
    for m in masters:
        w.writerow(
            [
                m.get("display_name", ""),
                m.get("revenue", ""),
                m.get("completed_bookings", ""),
                m.get("avg_check", ""),
                m.get("utilization_pct", ""),
            ]
        )
    w.writerow([])
    w.writerow(["Top services"])
    w.writerow(["name", "revenue", "completed_bookings"])
    for s in services_top:
        name_i18n = s.get("name_i18n") or {}
        name = name_i18n.get("en") or next(iter(name_i18n.values()), "")
        w.writerow([name, s.get("revenue", ""), s.get("completed_bookings", "")])
    return buf.getvalue().encode("utf-8-sig")


def stats_snapshot_to_pdf(
    *,
    period_label: str,
    kpi: dict[str, Any],
    masters: list[dict[str, Any]],
    services_top: list[dict[str, Any]],
    currency: str = "EUR",
) -> bytes:
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 10, "Statistics report", ln=True)
    pdf.set_font("Helvetica", size=10)
    pdf.cell(0, 8, period_label, ln=True)
    pdf.ln(3)

    pdf.set_font("Helvetica", "B", 11)
    pdf.cell(0, 7, "KPI", ln=True)
    pdf.set_font("Helvetica", size=9)
    for label, key in [
        ("Revenue", "revenue"),
        ("Completed bookings", "completed_bookings"),
        ("Avg check", "avg_check"),
        ("New clients", "new_clients_count"),
        ("Cancelled", "cancelled_bookings_count"),
    ]:
        val = kpi.get(key, "")
        pdf.cell(0, 6, f"{label}: {val} {currency if 'revenue' in key or 'check' in key else ''}".strip(), ln=True)
    pdf.ln(3)

    pdf.set_font("Helvetica", "B", 11)
    pdf.cell(0, 7, "Masters", ln=True)
    pdf.set_font("Helvetica", "B", 9)
    for h, w in [("Master", 50), ("Revenue", 30), ("Done", 18), ("Avg", 30), ("Util %", 22)]:
        pdf.cell(w, 7, h, border=1)
    pdf.ln()
    pdf.set_font("Helvetica", size=9)
    for m in masters:
        pdf.cell(50, 7, str(m.get("display_name", ""))[:28], border=1)
        pdf.cell(30, 7, str(m.get("revenue", ""))[:12], border=1)
        pdf.cell(18, 7, str(m.get("completed_bookings", ""))[:6], border=1)
        pdf.cell(30, 7, str(m.get("avg_check", ""))[:12], border=1)
        pdf.cell(22, 7, str(m.get("utilization_pct", ""))[:8], border=1)
        pdf.ln()

    pdf.ln(3)
    pdf.set_font("Helvetica", "B", 11)
    pdf.cell(0, 7, "Top services", ln=True)
    pdf.set_font("Helvetica", "B", 9)
    for h, w in [("Service", 90), ("Revenue", 30), ("Bookings", 25)]:
        pdf.cell(w, 7, h, border=1)
    pdf.ln()
    pdf.set_font("Helvetica", size=9)
    for s in services_top:
        name_i18n = s.get("name_i18n") or {}
        name = name_i18n.get("en") or next(iter(name_i18n.values()), "")
        pdf.cell(90, 7, str(name)[:48], border=1)
        pdf.cell(30, 7, str(s.get("revenue", ""))[:12], border=1)
        pdf.cell(25, 7, str(s.get("completed_bookings", ""))[:8], border=1)
        pdf.ln()

    out = pdf.output(dest="S")
    if isinstance(out, str):
        return out.encode("latin-1")
    return bytes(out)
