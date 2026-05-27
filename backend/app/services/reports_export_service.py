"""Экспорт отчётов в XLSX / PDF."""

from __future__ import annotations

from io import BytesIO
from typing import Any

from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Font


def pnl_to_xlsx(pnl: dict[str, Any], *, salon_name: str, period_label: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "P&L"
    ws.append([salon_name, "P&L Report", period_label])
    ws.append([])
    ws.append(["Revenue", "Amount"])
    rev = pnl.get("revenue", {})
    ws.append(["Services", float(rev.get("services", 0))])
    ws.append(["Products", float(rev.get("products", 0))])
    ws.append(["Total revenue", float(rev.get("total", 0))])
    ws.append([])
    ws.append(["Expenses", "Amount"])
    exp = pnl.get("expenses", {})
    for key in (
        "salaries",
        "rent",
        "utilities",
        "supplies",
        "advertising",
        "equipment",
        "taxes",
        "software",
        "training",
        "other",
    ):
        ws.append([key.replace("_", " ").title(), float(exp.get(key, 0))])
    ws.append(["Total expenses", float(exp.get("total", 0))])
    ws.append([])
    ws.append(["Gross profit", float(pnl.get("gross_profit", 0))])
    ws.append(["Margin %", float(pnl.get("profit_margin_percent", 0))])
    for row in ws.iter_rows(min_row=3, max_row=3):
        for c in row:
            c.font = Font(bold=True)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def cash_to_xlsx(report: dict[str, Any], *, period_label: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Cash"
    ws.append(["Cash report", period_label])
    summary = report.get("summary", {})
    ws.append(["Total income", float(summary.get("total_income", 0))])
    ws.append(["Cash", float(summary.get("income_cash", 0))])
    ws.append(["Card", float(summary.get("income_card", 0))])
    ws.append(["Expenses", float(summary.get("total_expenses", 0))])
    ws.append(["Balance", float(summary.get("balance", 0))])
    ws.append([])
    ws.append(["Date", "Type", "Source", "Amount", "Cash", "Card", "Description"])
    for day in report.get("by_day", []):
        for tx in day.get("transactions", []):
            ws.append(
                [
                    day.get("date"),
                    tx.get("type"),
                    tx.get("source"),
                    float(tx.get("amount", 0)),
                    float(tx.get("cash") or 0) if tx.get("cash") is not None else "",
                    float(tx.get("card") or 0) if tx.get("card") is not None else "",
                    tx.get("description", ""),
                ]
            )
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def salaries_to_pdf(
    salaries: dict[str, Any], *, salon_name: str, period_label: str
) -> bytes:
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 10, salon_name, ln=True)
    pdf.set_font("Helvetica", size=10)
    pdf.cell(0, 8, f"Salary report — {period_label}", ln=True)
    pdf.ln(4)
    for m in salaries.get("masters", []):
        pdf.set_font("Helvetica", "B", 11)
        pdf.cell(0, 8, str(m.get("name", "")), ln=True)
        pdf.set_font("Helvetica", size=9)
        pdf.cell(
            0,
            6,
            f"Revenue: {m.get('revenue')} | Bookings: {m.get('bookings_count')} | Salary: {m.get('calculated_salary')}",
            ln=True,
        )
        pdf.ln(2)
    pdf.set_font("Helvetica", "B", 10)
    pdf.cell(0, 8, f"Total: {salaries.get('total_calculated')}", ln=True)
    out = pdf.output(dest="S")
    if isinstance(out, str):
        return out.encode("latin-1", errors="replace")
    return bytes(out)
